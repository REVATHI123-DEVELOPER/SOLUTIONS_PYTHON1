# pip install pandas openpyxl

import argparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers

from db.db_reports import get_recommended_workitems_per_release

# Predefined repo_id mappings
repo_id_mapping = {
    "b38f60fc-9175-48e4-97ae-ddf75e4e4bb6": "BE",
    "5e42b435-2f86-41b5-adba-ba805be3d358": "Web",
    "ea77b4af-3f3f-4e32-bf9d-3c08e1d0d9ea": "Mobile",
}


def run(release):
    # Fetch the rows from the database based on the release
    rows = get_recommended_workitems_per_release(release)

    # Define column names manually
    columns = [
        "PR",
        "Repo",
        "Associated WorkItem",
        "Associated WorkItem Feature",
        "Associated WorkItem Target Release",
        "Recommended Workitem",
        "Recommended WorkItem Priority",
        "Recommended WorkItem Severity",
        "Recommended WorkItem Feature",
        "Recommended Date",
        "Confidence Score",
    ]

    # Convert the rows to a pandas DataFrame
    df = pd.DataFrame(rows, columns=columns)

    # Map the 'Repo' to human-readable names (BE, Web, Mobile)
    df["Repo"] = df["Repo"].map(repo_id_mapping).fillna(df["Repo"])

    # Convert all values in the DataFrame to strings using apply
    df = df.apply(lambda x: x.astype(str))

    unique_recommended_workitems = []
    unique_recommended_workitems_count = []
    for index, row in df.iterrows():
        if row["Recommended Workitem"] not in unique_recommended_workitems:
            unique_recommended_workitems.append(row["Recommended Workitem"])
        unique_recommended_workitems_count.append(len(unique_recommended_workitems))

    df["Unique Count"] = unique_recommended_workitems_count

    # Write the DataFrame to an Excel file in the current directory
    file_name = f"recommended_workitems_{release}.xlsx"
    df.to_excel(file_name, index=False, header=True)

    # Now format the cells as text in Excel using openpyxl
    wb = load_workbook(file_name)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column - 1):
        for cell in row:
            cell.number_format = numbers.FORMAT_TEXT
            cell.value = str(cell.value)  # Ensure value is string type

    wb.save(file_name)


if __name__ == "__main__":
    # Set up argument parsing
    parser = argparse.ArgumentParser(
        description="Fetch recommended workitems for a specific release."
    )
    parser.add_argument(
        "--release",
        type=str,
        help="The release name to fetch recommended workitems for.",
    )

    # Parse the arguments
    args = parser.parse_args()

    # Run with the provided release
    run(args.release)
